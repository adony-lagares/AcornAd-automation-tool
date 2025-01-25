import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
import tkinter as tk
from tkinter import filedialog, messagebox
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter


def formatar_planilha_saida(caminho_arquivo):
    """
    Aplica formatações à planilha de saída para melhorar a legibilidade.
    Ajusta larguras das colunas automaticamente e formata os cabeçalhos.

    Args:
        caminho_arquivo (str): Caminho da planilha Excel a ser formatada.
    """
    workbook = load_workbook(caminho_arquivo)
    sheet = workbook.active

    # Estiliza os cabeçalhos
    estilo_cabecalho = Font(bold=True, size=12)
    for celula in sheet[1]:
        celula.font = estilo_cabecalho
        celula.alignment = Alignment(horizontal="center", vertical="center")

    # Ajusta a largura das colunas com base no conteúdo
    for coluna in sheet.columns:
        maximo = max(len(str(celula.value)) for celula in coluna if celula.value is not None)
        largura_ajustada = maximo + 2  # Adiciona espaçamento extra
        sheet.column_dimensions[get_column_letter(coluna[0].column)].width = largura_ajustada

    workbook.save(caminho_arquivo)


def buscar_no_youtube(palavra_chave, driver):
    """
    Realiza uma busca no YouTube pela palavra-chave fornecida e extrai detalhes do vídeo.

    Args:
        palavra_chave (str): Palavra-chave para a busca.
        driver (WebDriver): Instância do Selenium WebDriver.

    Returns:
        dict: Um dicionário com informações do vídeo (título, link, link do canal).
    """
    try:
        driver.get("https://www.youtube.com")
        campo_busca = driver.find_element(By.NAME, "search_query")
        campo_busca.send_keys(palavra_chave)
        campo_busca.submit()

        driver.implicitly_wait(5)  # Aguarda resultados carregarem

        # Captura o primeiro resultado válido
        videos = driver.find_elements(By.ID, "video-title")
        for video in videos:
            link = video.get_attribute("href")
            if link is not None:
                titulo = video.get_attribute("title")
                try:
                    elemento_canal = video.find_element(By.XPATH, "./../../..//yt-formatted-string[@id='text']//a")
                    link_canal = elemento_canal.get_attribute("href")
                except Exception:
                    link_canal = "N/A"

                return {
                    "Palavra-chave": palavra_chave,
                    "Título": titulo,
                    "Link": link,
                    "Link do Canal": link_canal
                }

        return {
            "Palavra-chave": palavra_chave,
            "Título": "N/A",
            "Link": "N/A",
            "Link do Canal": "N/A"
        }
    except Exception as e:
        return {
            "Palavra-chave": palavra_chave,
            "Título": "Erro na busca",
            "Link": "N/A",
            "Link do Canal": f"Erro: {e}"
        }


def processar_arquivo_entrada(arquivo_entrada, arquivo_saida, widget_log):
    """
    Processa o arquivo de entrada (Excel) realizando buscas no YouTube para cada palavra-chave
    e salva os resultados no arquivo de saída.

    Args:
        arquivo_entrada (str): Caminho do arquivo Excel com as palavras-chave.
        arquivo_saida (str): Caminho do arquivo Excel de saída.
        widget_log (Listbox): Widget de log para exibir mensagens de progresso.
    """
    try:
        # Configuração do WebDriver
        opcoes_chrome = Options()
        opcoes_chrome.add_argument("--disable-gpu")
        opcoes_chrome.add_argument("--headless")
        driver_service = Service("./chromedriver.exe")
        driver = webdriver.Chrome(service=driver_service, options=opcoes_chrome)

        # Leitura do arquivo de entrada
        palavras_chave = pd.read_excel(arquivo_entrada)
        resultados = []

        for _, linha in palavras_chave.iterrows():
            palavra_chave = linha["Keyword"]
            widget_log.insert(tk.END, f"Buscando: {palavra_chave}")
            widget_log.see(tk.END)
            widget_log.update()

            resultado = buscar_no_youtube(palavra_chave, driver)
            resultados.append(resultado)

        # Salva os resultados no arquivo de saída
        pd.DataFrame(resultados).to_excel(arquivo_saida, index=False)
        formatar_planilha_saida(arquivo_saida)

        widget_log.insert(tk.END, f"Resultados salvos em: {arquivo_saida}")
        widget_log.see(tk.END)
        driver.quit()
    except Exception as e:
        widget_log.insert(tk.END, f"Erro ao processar: {e}")
        widget_log.see(tk.END)


def selecionar_arquivo(widget_entrada):
    """
    Abre uma janela para o usuário selecionar o arquivo de entrada e insere o caminho no widget.

    Args:
        widget_entrada (Entry): Widget de entrada de texto para exibir o caminho do arquivo.
    """
    caminho = filedialog.askopenfilename(filetypes=[("Arquivos Excel", "*.xlsx")])
    if caminho:
        widget_entrada.delete(0, tk.END)
        widget_entrada.insert(0, caminho)


def iniciar_processamento(widget_entrada, widget_log, widget_saida):
    """
    Inicia o processamento do arquivo selecionado, realizando buscas no YouTube
    e gerando a planilha de saída.

    Args:
        widget_entrada (Entry): Widget de entrada de texto com o caminho do arquivo.
        widget_log (Listbox): Widget de log para exibir mensagens de progresso.
        widget_saida (Label): Widget de rótulo para exibir o caminho do arquivo de saída.
    """
    arquivo_entrada = widget_entrada.get()
    if not arquivo_entrada or not os.path.exists(arquivo_entrada):
        messagebox.showerror("Erro", "Selecione um arquivo válido.")
        return

    arquivo_saida = "search_output.xlsx"
    widget_saida.config(text=f"Arquivo gerado: {arquivo_saida}")
    widget_log.delete(0, tk.END)
    widget_log.insert(tk.END, "Iniciando o processamento...")
    widget_log.see(tk.END)
    processar_arquivo_entrada(arquivo_entrada, arquivo_saida, widget_log)


def abrir_arquivo_saida(widget_saida):
    """
    Abre o arquivo de saída no programa padrão do sistema.

    Args:
        widget_saida (Label): Widget que contém o caminho do arquivo de saída.
    """
    caminho = widget_saida.cget("text").replace("Arquivo gerado: ", "")
    if os.path.exists(caminho):
        os.startfile(caminho)
    else:
        messagebox.showerror("Erro", "Arquivo de saída não encontrado.")


def criar_interface():
    """
    Cria a interface gráfica da aplicação utilizando Tkinter.
    """
    root = tk.Tk()
    root.title("Automação de Busca no YouTube")
    root.geometry("650x500")
    root.configure(bg="#f4f4f4")

    tk.Label(root, text="Automação de Busca no YouTube", font=("Helvetica", 16, "bold"), bg="#f4f4f4").pack(pady=10)

    frame_entrada = tk.Frame(root, bg="#f4f4f4")
    frame_entrada.pack(pady=10)

    tk.Label(frame_entrada, text="Arquivo de entrada (Excel):", bg="#f4f4f4").pack(side=tk.LEFT, padx=5)
    widget_entrada = tk.Entry(frame_entrada, width=40)
    widget_entrada.pack(side=tk.LEFT, padx=5)
    tk.Button(frame_entrada, text="Procurar", command=lambda: selecionar_arquivo(widget_entrada), bg="#4CAF50", fg="white").pack(side=tk.LEFT, padx=5)

    tk.Button(root, text="Iniciar Busca", command=lambda: iniciar_processamento(widget_entrada, log_box, label_saida), bg="#4CAF50", fg="white").pack(pady=10)

    label_saida = tk.Label(root, text="", bg="#f4f4f4", font=("Helvetica", 10))
    label_saida.pack(pady=5)

    tk.Button(root, text="Abrir Resultado", command=lambda: abrir_arquivo_saida(label_saida), bg="#2196F3", fg="white").pack(pady=10)

    tk.Label(root, text="Log de Execução:", bg="#f4f4f4").pack(pady=10)
    log_box = tk.Listbox(root, width=60, height=10)
    log_box.pack(pady=5)

    tk.Label(root, text="Desenvolvido por Ádony Lagares", font=("Helvetica", 10), bg="#f4f4f4", fg="#666").pack(pady=10)

    root.mainloop()


if __name__ == "__main__":
    criar_interface()
